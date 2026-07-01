import io
import unittest

import pandas as pd
from openpyxl import load_workbook

from reconciliation import (
    SPECIAL_REMARK_KEYWORDS,
    VALIDATION_FOUND,
    VALIDATION_NOT_FOUND,
    VALIDATION_REVIEW,
    build_excel_report,
    reconcile_dat_vs_stream,
    special_stream_remark_keyword,
    validate_required_columns,
)


class ReconciliationTests(unittest.TestCase):
    def setUp(self):
        self.dep = pd.DataFrame(
            {
                "'callsign'": ["'LNI201'", "'LNI201'", "'CTV911'"],
                "'adep'": ["'WIMM'", "'WIMM'", "'WIMM'"],
                "'ades'": ["'WIII'", "'WIII'", "'WSSS'"],
                "'eobd'": ["'260523'", "'260523'", "'260523'"],
                "'atd'": ["'2026-05-23 22:19:00'", "'2026-05-23 22:19:00'", "'23:01'"],
                "'departureRunway'": ["'23'", "'23'", "'23'"],
                "'departureGate'": ["'30'", "'30'", "'27'"],
                "'register'": ["'PKLHO'", "'PKLHO'", "'PKGTG'"],
            }
        )
        self.arr = pd.DataFrame(
            {
                "'callsign'": ["'LNI200'", "'WON1265'"],
                "'adep'": ["'WIII'", "'WIMB'"],
                "'ades'": ["'WIMM'", "'WIMM'"],
                "'eobd'": ["'260523'", "'260523'"],
                "'ata'": ["'2026-05-24 00:09:00'", "'2026-05-24 00:28:00'"],
                "'arrivalRunway'": ["'23'", "'23'"],
                "'arrivalGate'": ["'30'", "'11'"],
                "'register'": ["'PKLHO'", "'PKWHT'"],
            }
        )
        self.stream = pd.DataFrame(
            {
                "DATE OF FLIGHT": ["2026-05-23", "2026-05-23", "2026-05-23"],
                "FLIGHT NUMBER": ["LNI201", "LNI200", "EXTRA1"],
                "AERODROME": ["WIMM", "WIII", "WIMM"],
                "TO FROM": ["WIII", "WIMM", "WIII"],
                "AC REGISTER": ["PKLHO", "PKLHO", "PKXXX"],
                "D/A/L/O": ["D", "A", "D"],
                "RWY": ["23", "23", "05"],
                "PARKING": ["30", "30", "10"],
                "ATD": ["2026-05-23 22:19", "", "09:00"],
                "ATA": ["", "2026-05-24 00:09", ""],
            }
        )

    def test_reconciliation_counts_and_fields(self):
        result = reconcile_dat_vs_stream(self.dep, self.arr, self.stream)
        summary = result["summary"]
        self.assertEqual(summary["total_dat_dep"], 3)
        self.assertEqual(summary["total_dat_arr"], 2)
        self.assertEqual(summary["total_dat_combined"], 5)
        self.assertEqual(summary["duplicate_dat"], 1)
        self.assertEqual(summary["matched"], 2)
        self.assertEqual(summary["missing_in_stream"], 2)
        self.assertEqual(summary["extra_in_stream"], 1)
        self.assertAlmostEqual(summary["accuracy_percentage"], 50.0)

        matched_dep = result["matched"].loc[
            result["matched"]["FLIGHT NUMBER"] == "LNI201"
        ].iloc[0]
        self.assertEqual(matched_dep["DATE OF FLIGHT"], "2026-05-23")
        self.assertEqual(matched_dep["DEPARTURE GATE"], "30")
        self.assertEqual(matched_dep["DEPARTURE RUNWAY"], "23")
        self.assertEqual(matched_dep["ATD"], "22:19")

    def test_excel_report_has_required_sheets(self):
        result = reconcile_dat_vs_stream(self.dep, self.arr, self.stream)
        workbook = load_workbook(io.BytesIO(build_excel_report(result)), read_only=True)
        self.assertEqual(
            workbook.sheetnames,
            [
                "Summary",
                "Validasi",
                "Missing in Stream",
                "Matched",
                "Need Review",
                "Extra in Stream",
                "Duplicate DAT",
                "Audit Detail",
                "Excluded Non-Billable",
            ],
        )
        self.assertEqual(workbook["Summary"]["A1"].value, "FINDER — DAT vs STREAM Reconciliation Report")
        missing_headers = [
            cell.value for cell in workbook["Missing in Stream"][4]
        ]
        self.assertEqual(
            missing_headers,
            [
                "DATE OF FLIGHT",
                "ACTUAL MOVEMENT DATE",
                "FLIGHT NUMBER",
                "AERODROME",
                "TO FROM",
                "AC REGISTER",
                "ATD",
                "ATA",
                "D/A/L/O",
                "ARRIVAL GATE",
                "DEPARTURE GATE",
                "DEPARTURE RUNWAY",
                "ARRIVAL RUNWAY",
                "VALIDASI",
                "STATUS",
                "MATCH_REASON",
                "STREAM STATUS FLIGHT",
                "DAT MOVEMENT DATETIME",
                "STREAM MOVEMENT DATETIME",
                "TIME DIFFERENCE MINUTES",
                "STREAM REMARK",
                "STREAM SPECIAL REMARK FLAG",
                "ROUTE MATCH IGNORED",
                "SPECIAL REMARK KEYWORD FOUND",
                "DAT_RECOVERY_USED",
                "DAT_RECOVERY_REASON",
                "DAT_RECOVERY_SOURCE_DATE",
                "DAT_RECOVERY_SOURCE_ROW",
                "ORIGINAL_DAT_DATE",
                "RECOVERED_DAT_DATE",
                "USED_FOR_RECOVERY",
                "STREAM MATCH DATE USED",
                "STREAM MATCH MODE",
            ],
        )

    def test_validation_reports_missing_stream_movement(self):
        mapping = {"flight": "f", "eobd": "d", "adep": "o", "ades": "x"}
        issues = validate_required_columns(mapping, mapping, mapping)
        self.assertEqual(len(issues), 1)
        self.assertIn("D/A/L/O", issues[0])

    def test_sia990_selects_complete_arrival_record(self):
        dep = pd.DataFrame(columns=["callsign", "adep", "ades", "eobd"])
        arr = pd.DataFrame(
            {
                "callsign": ["SIA990", "SIA990"],
                "adep": ["WSSS", "WSSS"],
                "ades": ["WIMM", "WIMM"],
                "eobd": ["260529", "260529"],
                "atd": ["", "23:44"],
                "ata": ["", "00:48"],
                "register": ["", "9VMBT"],
                "arrivalGate": ["", "13"],
                "arrivalRunway": ["", "23"],
                "timeStamp": ["2026-05-29 23:00", "2026-05-30 01:00"],
                "messageNum": ["1", "2"],
            }
        )
        stream = pd.DataFrame(
            columns=[
                "DATE OF FLIGHT",
                "FLIGHT NUMBER",
                "AERODROME",
                "TO FROM",
                "D/A/L/O",
            ]
        )

        result = reconcile_dat_vs_stream(dep, arr, stream)
        selected = result["missing"].iloc[0]
        self.assertEqual(selected["ATD"], "23:44")
        self.assertEqual(selected["ATA"], "00:48")
        self.assertEqual(selected["AC REGISTER"], "9VMBT")
        self.assertEqual(selected["ARRIVAL GATE"], "13")
        self.assertEqual(selected["ARRIVAL RUNWAY"], "23")
        self.assertEqual(selected["DAT MOVEMENT DATETIME"], "2026-05-30 00:48")
        self.assertEqual(selected["ACTUAL MOVEMENT DATE"], "2026-05-30")
        self.assertEqual(len(result["duplicates"]), 1)
        duplicate = result["duplicates"].iloc[0]
        self.assertFalse(bool(duplicate["SELECTED_RECORD_FLAG"]))
        self.assertEqual(int(duplicate["COMPLETENESS_SCORE"]), 0)
        self.assertFalse(bool(duplicate["HAS_MOVEMENT_TIME"]))

    def test_2ambo_complete_record_is_missing_for_other_and_wrong_date(self):
        dep = pd.DataFrame(columns=["callsign", "adep", "ades", "eobd"])
        arr = pd.DataFrame(
            {
                "callsign": ["2AMBO", "2AMBO"],
                "adep": ["ZGGG", "ZGGG"],
                "ades": ["WIMM", "WIMM"],
                "eobd": ["260620", "260620"],
                "atd": ["", "2026-06-20 10:36"],
                "ata": ["", "2026-06-20 14:08"],
                "register": ["", "2AMBO"],
                "arrivalGate": ["", "01"],
                "arrivalRunway": ["", "23"],
                "timeStamp": ["2026-06-20 11:00", "2026-06-20 15:00"],
                "messageNum": ["10", "11"],
            }
        )
        stream = pd.DataFrame(
            {
                "DATE OF FLIGHT": ["2026-06-20"],
                "FLIGHT NUMBER": ["2AMBO"],
                "AERODROME": ["ZGGG"],
                "TO FROM": ["WIMM"],
                "D/A/L/O": ["A"],
                "ATA": ["2026-06-19 14:08"],
                "STATUS FLIGHT": ["OTHER"],
            }
        )

        result = reconcile_dat_vs_stream(dep, arr, stream)
        self.assertEqual(result["summary"]["matched"], 0)
        self.assertEqual(result["summary"]["missing_in_stream"], 1)
        missing = result["missing"].iloc[0]
        self.assertEqual(missing["ATD"], "10:36")
        self.assertEqual(missing["ATA"], "14:08")
        self.assertEqual(missing["AC REGISTER"], "2AMBO")
        self.assertEqual(missing["ARRIVAL GATE"], "01")
        self.assertEqual(missing["ARRIVAL RUNWAY"], "23")
        self.assertIn("STREAM STATUS OTHER", missing["MATCH_REASON"])
        self.assertIn("STREAM INVALID ATA DATE", missing["MATCH_REASON"])
        self.assertEqual(len(result["duplicates"]), 1)

    def test_general_duplicate_prefers_movement_time_then_completeness(self):
        dep = pd.DataFrame(
            {
                "callsign": ["ABC123", "ABC123", "ABC123"],
                "adep": ["WIMM"] * 3,
                "ades": ["WIII"] * 3,
                "eobd": ["260620"] * 3,
                "atd": ["", "08:00", "08:00"],
                "register": ["PKAAA", "", "PKBEST"],
                "departureGate": ["10", "", "12"],
                "departureRunway": ["23", "", "23"],
                "timeStamp": [
                    "2026-06-20 09:00",
                    "2026-06-20 08:00",
                    "2026-06-20 07:00",
                ],
                "messageNum": ["30", "20", "10"],
            }
        )
        arr = pd.DataFrame(columns=["callsign", "adep", "ades", "eobd"])
        stream = pd.DataFrame(
            columns=[
                "DATE OF FLIGHT",
                "FLIGHT NUMBER",
                "AERODROME",
                "TO FROM",
                "D/A/L/O",
            ]
        )
        result = reconcile_dat_vs_stream(dep, arr, stream)
        selected = result["dat_unique"].iloc[0]
        self.assertEqual(selected["AC REGISTER"], "PKBEST")
        self.assertEqual(selected["DEPARTURE GATE"], "12")
        self.assertEqual(len(result["duplicates"]), 2)

    def test_time_tolerance_and_date_validation(self):
        dep = pd.DataFrame(
            {
                "callsign": ["TOL30", "TOL31", "TOL121", "DATEBAD"],
                "adep": ["WIMM"] * 4,
                "ades": ["WIII"] * 4,
                "eobd": ["260620"] * 4,
                "atd": ["10:00", "11:00", "12:00", "13:00"],
            }
        )
        arr = pd.DataFrame(columns=["callsign", "adep", "ades", "eobd"])
        stream = pd.DataFrame(
            {
                "DATE OF FLIGHT": ["2026-06-20"] * 4,
                "FLIGHT NUMBER": ["TOL30", "TOL31", "TOL121", "DATEBAD"],
                "AERODROME": ["WIMM"] * 4,
                "TO FROM": ["WIII"] * 4,
                "D/A/L/O": ["D"] * 4,
                "ATD": [
                    "2026-06-20 10:30",
                    "2026-06-20 11:31",
                    "2026-06-20 14:01",
                    "2026-06-19 13:00",
                ],
                "STATUS FLIGHT": ["ACTIVE"] * 4,
            }
        )
        result = reconcile_dat_vs_stream(
            dep, arr, stream, time_tolerance_minutes=30
        )
        self.assertEqual(set(result["matched"]["FLIGHT NUMBER"]), {"TOL30"})
        self.assertEqual(set(result["need_review"]["FLIGHT NUMBER"]), {"TOL31"})
        self.assertEqual(
            set(result["missing"]["FLIGHT NUMBER"]), {"TOL121", "DATEBAD"}
        )
        reasons = result["missing"].set_index("FLIGHT NUMBER")["MATCH_REASON"]
        self.assertEqual(reasons["TOL121"], "STREAM TIME MISMATCH")
        self.assertIn("STREAM INVALID ATD DATE", reasons["DATEBAD"])

    def test_validation_only_contains_dat_without_any_stream_candidate(self):
        dep = pd.DataFrame(
            {
                "callsign": ["FOUND1", "ABSENT1", "OTHER1"],
                "adep": ["WIMM"] * 3,
                "ades": ["WIII"] * 3,
                "eobd": ["260624"] * 3,
                "atd": ["10:00", "11:00", "12:00"],
                "register": ["PKAAA", "PKBBB", "PKCCC"],
            }
        )
        arr = pd.DataFrame(columns=["callsign", "adep", "ades", "eobd"])
        stream = pd.DataFrame(
            {
                "DATE OF FLIGHT": ["2026-06-24", "2026-06-24"],
                "FLIGHT NUMBER": ["FOUND1", "OTHER1"],
                "AERODROME": ["WIMM", "WIMM"],
                "TO FROM": ["WIII", "WIII"],
                "AC REGISTER": ["PKAAA", "PKCCC"],
                "D/A/L/O": ["D", "D"],
                "ATD": ["2026-06-24 10:05", "2026-06-24 12:00"],
                "STATUS FLIGHT": ["REGULER", "OTHER"],
            }
        )

        result = reconcile_dat_vs_stream(dep, arr, stream)

        self.assertEqual(
            result["matched"].iloc[0]["VALIDASI"], VALIDATION_FOUND
        )
        missing_validation = result["missing"].set_index("FLIGHT NUMBER")[
            "VALIDASI"
        ]
        self.assertEqual(missing_validation["ABSENT1"], VALIDATION_NOT_FOUND)
        self.assertEqual(missing_validation["OTHER1"], VALIDATION_REVIEW)
        self.assertEqual(set(result["validasi"]["FLIGHT NUMBER"]), {"ABSENT1"})
        self.assertEqual(result["summary"]["total_ada_di_stream"], 1)
        self.assertEqual(result["summary"]["total_validasi"], 1)
        self.assertEqual(result["summary"]["total_perlu_review_stream"], 1)
        self.assertLessEqual(
            result["summary"]["total_validasi"],
            result["summary"]["missing_in_stream"],
        )

        workbook = load_workbook(
            io.BytesIO(build_excel_report(result)), read_only=True
        )
        validation_sheet = workbook["Validasi"]
        headers = [cell.value for cell in validation_sheet[4]]
        validation_column = headers.index("VALIDASI") + 1
        self.assertEqual(validation_sheet.max_row, 5)
        self.assertEqual(
            validation_sheet.cell(5, validation_column).value,
            VALIDATION_NOT_FOUND,
        )

    def test_actual_movement_index_matches_when_stream_flight_date_is_unrelated(self):
        dep = pd.DataFrame(
            {
                "callsign": ["ACTD01"],
                "adep": ["WIMM"],
                "ades": ["WIII"],
                "eobd": ["260624"],
                "atd": ["2026-06-25 00:10"],
                "register": ["PKDEP"],
            }
        )
        arr = pd.DataFrame(
            {
                "callsign": ["ACTA01"],
                "adep": ["WIII"],
                "ades": ["WIMM"],
                "eobd": ["260624"],
                "ata": ["2026-06-25 02:20"],
                "register": ["PKARR"],
            }
        )
        stream = pd.DataFrame(
            {
                "DATE OF FLIGHT": ["2026-06-23", "2026-06-23"],
                "FLIGHT NUMBER": ["ACTD01", "ACTA01"],
                "AERODROME": ["WIMM", "WIII"],
                "TO FROM": ["WIII", "WIMM"],
                "AC REGISTER": ["PKDEP", "PKARR"],
                "D/A/L/O": ["D", "A"],
                "ATD": ["2026-06-25 00:15", ""],
                "ATA": ["", "2026-06-25 02:25"],
                "STATUS FLIGHT": ["REGULER", "REGULER"],
            }
        )

        result = reconcile_dat_vs_stream(dep, arr, stream)

        self.assertEqual(set(result["matched"]["FLIGHT NUMBER"]), {"ACTD01", "ACTA01"})
        self.assertTrue(result["missing"].empty)
        self.assertTrue(result["validasi"].empty)
        self.assertEqual(
            set(result["matched"]["STREAM MATCH MODE"]),
            {"ACTUAL MOVEMENT DATE MATCH"},
        )
        self.assertEqual(
            set(result["matched"]["VALIDASI"]), {VALIDATION_FOUND}
        )

    def test_special_remark_ignores_route_and_is_excluded_from_validation(self):
        dep = pd.DataFrame(
            {
                "callsign": ["ABC123"],
                "adep": ["WIMM"],
                "ades": ["WIII"],
                "eobd": ["260525"],
                "atd": ["06:14"],
                "register": ["PKABC"],
            }
        )
        arr = pd.DataFrame(columns=["callsign", "adep", "ades", "eobd"])
        stream = pd.DataFrame(
            {
                "DATE OF FLIGHT": ["2026-05-25"],
                "FLIGHT NUMBER": ["ABC123"],
                "AERODROME": ["WIDD"],
                "TO FROM": ["WADD"],
                "AC REGISTER": ["PKABC"],
                "D/A/L/O": ["D"],
                "ATD": ["2026-05-25 06:16"],
                "STATUS FLIGHT": ["REGULER"],
                "REMARK": ["DIVERT"],
            }
        )

        result = reconcile_dat_vs_stream(dep, arr, stream)

        self.assertTrue(result["validasi"].empty)
        self.assertTrue(result["missing"].empty)
        self.assertEqual(len(result["need_review"]), 1)
        reviewed = result["need_review"].iloc[0]
        self.assertEqual(reviewed["VALIDASI"], VALIDATION_REVIEW)
        self.assertEqual(
            reviewed["MATCH_REASON"],
            "STREAM SPECIAL REMARK FOUND - ROUTE IGNORED FOR VALIDATION",
        )
        self.assertEqual(reviewed["STREAM MATCH MODE"], "SPECIAL REMARK MATCH")
        self.assertEqual(reviewed["STREAM REMARK"], "DIVERT")
        self.assertTrue(bool(reviewed["STREAM SPECIAL REMARK FLAG"]))
        self.assertTrue(bool(reviewed["ROUTE MATCH IGNORED"]))
        self.assertEqual(reviewed["SPECIAL REMARK KEYWORD FOUND"], "DIVERT")
        self.assertEqual(float(reviewed["TIME DIFFERENCE MINUTES"]), 2.0)
        self.assertEqual(result["summary"]["total_validasi"], 0)
        self.assertEqual(result["summary"]["total_perlu_review_stream"], 1)
        self.assertEqual(result["summary"]["extra_in_stream"], 0)

        workbook = load_workbook(
            io.BytesIO(build_excel_report(result)), read_only=True
        )
        self.assertEqual(workbook["Validasi"].max_row, 4)
        review_headers = [cell.value for cell in workbook["Need Review"][4]]
        for audit_column in (
            "STREAM REMARK",
            "STREAM SPECIAL REMARK FLAG",
            "STREAM MATCH MODE",
            "ROUTE MATCH IGNORED",
            "SPECIAL REMARK KEYWORD FOUND",
        ):
            self.assertIn(audit_column, review_headers)
        review_values = {
            header: workbook["Need Review"].cell(5, index + 1).value
            for index, header in enumerate(review_headers)
        }
        self.assertEqual(review_values["STREAM REMARK"], "DIVERT")
        self.assertTrue(review_values["STREAM SPECIAL REMARK FLAG"])
        self.assertTrue(review_values["ROUTE MATCH IGNORED"])
        self.assertEqual(
            review_values["SPECIAL REMARK KEYWORD FOUND"], "DIVERT"
        )

    def test_all_special_remark_keywords_are_detected(self):
        for keyword in SPECIAL_REMARK_KEYWORDS:
            with self.subTest(keyword=keyword):
                row = pd.Series({"NOTES": f"OPERATION {keyword} CONFIRMED"})
                self.assertEqual(special_stream_remark_keyword(row), keyword)

    def test_special_remark_route_fallback_requires_register_and_tolerance(self):
        dep = pd.DataFrame(
            {
                "callsign": ["REGFAIL", "TIMEFAIL"],
                "adep": ["WIMM", "WIMM"],
                "ades": ["WIII", "WIII"],
                "eobd": ["260525", "260525"],
                "atd": ["06:14", "07:00"],
                "register": ["PKAAA", "PKBBB"],
            }
        )
        arr = pd.DataFrame(columns=["callsign", "adep", "ades", "eobd"])
        stream = pd.DataFrame(
            {
                "DATE OF FLIGHT": ["2026-05-25", "2026-05-25"],
                "FLIGHT NUMBER": ["REGFAIL", "TIMEFAIL"],
                "AERODROME": ["WIDD", "WIDD"],
                "TO FROM": ["WADD", "WADD"],
                "AC REGISTER": ["PKXXX", "PKBBB"],
                "D/A/L/O": ["D", "D"],
                "ATD": ["2026-05-25 06:16", "2026-05-25 08:00"],
                "REMARKS": ["RTB", "ALTERNATE"],
            }
        )

        result = reconcile_dat_vs_stream(dep, arr, stream)

        self.assertEqual(
            set(result["validasi"]["FLIGHT NUMBER"]),
            {"REGFAIL", "TIMEFAIL"},
        )
        self.assertTrue(result["need_review"].empty)
        self.assertEqual(result["summary"]["total_validasi"], 2)

    def test_special_remark_arrival_accepts_adjacent_date_and_normal_wins(self):
        dep = pd.DataFrame(columns=["callsign", "adep", "ades", "eobd"])
        arr = pd.DataFrame(
            {
                "callsign": ["ARRSP1", "NORMAL1"],
                "adep": ["WIII", "WIII"],
                "ades": ["WIMM", "WIMM"],
                "eobd": ["260525", "260525"],
                "ata": ["2026-05-25 23:55", "2026-05-25 10:00"],
                "register": ["PKARR", "PKNRM"],
            }
        )
        stream = pd.DataFrame(
            {
                "DATE OF FLIGHT": [
                    "2026-05-26",
                    "2026-05-25",
                    "2026-05-25",
                ],
                "FLIGHT NUMBER": ["ARRSP1", "NORMAL1", "NORMAL1"],
                "AERODROME": ["WIDD", "WIII", "WIDD"],
                "TO FROM": ["WADD", "WIMM", "WADD"],
                "AC REGISTER": ["PKARR", "PKNRM", "PKNRM"],
                "D/A/L/O": ["A", "A", "A"],
                "ATA": [
                    "2026-05-26 00:10",
                    "2026-05-25 10:05",
                    "2026-05-25 10:00",
                ],
                "STATUS FLIGHT": ["RTB", "REGULER", "REGULER"],
                "NOTE": ["", "", "DIVERSION"],
            }
        )

        result = reconcile_dat_vs_stream(dep, arr, stream)

        special = result["need_review"].loc[
            result["need_review"]["FLIGHT NUMBER"].eq("ARRSP1")
        ].iloc[0]
        self.assertEqual(special["STREAM MATCH MODE"], "SPECIAL REMARK MATCH")
        self.assertEqual(special["SPECIAL REMARK KEYWORD FOUND"], "RTB")
        self.assertEqual(float(special["TIME DIFFERENCE MINUTES"]), 15.0)
        normal = result["matched"].loc[
            result["matched"]["FLIGHT NUMBER"].eq("NORMAL1")
        ].iloc[0]
        self.assertEqual(normal["VALIDASI"], VALIDATION_FOUND)
        self.assertEqual(
            normal["STREAM MATCH MODE"], "ACTUAL MOVEMENT DATE MATCH"
        )
        self.assertFalse(bool(normal["STREAM SPECIAL REMARK FLAG"]))

    def test_non_billable_flight_is_hard_excluded(self):
        dep = pd.DataFrame(
            {
                "callsign": ["TEST1", "LNI777"],
                "adep": ["WIMM", "WIMM"],
                "ades": ["WIII", "WIII"],
                "eobd": ["260620", "260620"],
                "atd": ["08:00", "09:00"],
            }
        )
        arr = pd.DataFrame(columns=["callsign", "adep", "ades", "eobd"])
        stream = pd.DataFrame(
            columns=[
                "DATE OF FLIGHT",
                "FLIGHT NUMBER",
                "AERODROME",
                "TO FROM",
                "D/A/L/O",
            ]
        )
        result = reconcile_dat_vs_stream(dep, arr, stream)
        self.assertEqual(set(result["missing_billing"]["FLIGHT NUMBER"]), {"LNI777"})
        self.assertTrue(result["missing_non_billable"].empty)
        self.assertEqual(result["summary"]["excluded_dat_non_billable"], 1)
        self.assertEqual(result["summary"]["total_dat_combined"], 1)
        self.assertEqual(
            set(result["excluded_non_billable"]["FLIGHT NUMBER"]), {"TEST1"}
        )
        for key in ("missing", "matched", "need_review", "extra", "duplicates", "audit_detail"):
            self.assertNotIn("TEST1", set(result[key]["FLIGHT NUMBER"]))

    def test_all_internal_keywords_are_excluded_from_dat_stream_and_outputs(self):
        keywords = [
            "LANDASAN", "RWYINS", "RWYINSP", "INSPCTN", "CAR", "VFR",
            "IFR", "TEST", "TEST1", "TEST2", "MPS",
        ]
        callsigns = [f"X{keyword}9" for keyword in keywords]
        dep = pd.DataFrame(
            {
                "callsign": callsigns + ["BILLABLE9"],
                "adep": ["WIMM"] * (len(callsigns) + 1),
                "ades": ["WIII"] * (len(callsigns) + 1),
                "eobd": ["260620"] * (len(callsigns) + 1),
                "atd": ["08:00"] * (len(callsigns) + 1),
            }
        )
        arr = pd.DataFrame(columns=["callsign", "adep", "ades", "eobd"])
        stream = pd.DataFrame(
            {
                "DATE OF FLIGHT": ["2026-06-20"] * len(callsigns),
                "FLIGHT NUMBER": callsigns,
                "AERODROME": ["WIMM"] * len(callsigns),
                "TO FROM": ["WIII"] * len(callsigns),
                "D/A/L/O": ["D"] * len(callsigns),
                "ATD": ["08:00"] * len(callsigns),
            }
        )
        result = reconcile_dat_vs_stream(dep, arr, stream)
        self.assertEqual(result["summary"]["excluded_dat_non_billable"], len(callsigns))
        self.assertEqual(result["summary"]["excluded_stream_non_billable"], len(callsigns))
        self.assertEqual(result["summary"]["total_dat_combined"], 1)
        self.assertEqual(result["summary"]["total_stream"], 0)
        for key in ("missing", "matched", "need_review", "extra", "duplicates", "audit_detail"):
            values = set(result[key]["FLIGHT NUMBER"])
            self.assertTrue(values.isdisjoint(callsigns))

    def test_ctv010_adjacent_date_midnight_recovery_precedes_deduplication(self):
        dep = pd.DataFrame(columns=["callsign", "adep", "ades", "eobd"])
        arr = pd.DataFrame(
            {
                "callsign": ["CTV010", "CTV010"],
                "adep": ["WIHH", "WIHH"],
                "ades": ["WIMM", "WIMM"],
                "eobd": ["260623", "260624"],
                "eobt": ["2340", "0045"],
                "atd": ["", "2026-06-24 00:54"],
                "ata": ["", "2026-06-24 02:52"],
                "register": ["PKGLM", "PKGLM"],
                "arrivalGate": ["", "27"],
                "arrivalRunway": ["23", "23"],
                "timeStamp": ["2026-06-24 01:49", "2026-06-24 04:52"],
                "messageNum": ["52538", "52582"],
            }
        )
        stream = pd.DataFrame(
            columns=["DATE OF FLIGHT", "FLIGHT NUMBER", "AERODROME", "TO FROM", "D/A/L/O"]
        )
        result = reconcile_dat_vs_stream(dep, arr, stream)
        recovered = result["missing"].iloc[0]
        self.assertEqual(recovered["DATE OF FLIGHT"], "2026-06-23")
        self.assertEqual(recovered["ACTUAL MOVEMENT DATE"], "2026-06-24")
        self.assertEqual(recovered["ATD"], "00:54")
        self.assertEqual(recovered["ATA"], "02:52")
        self.assertEqual(recovered["ARRIVAL GATE"], "27")
        self.assertEqual(recovered["ARRIVAL RUNWAY"], "23")
        self.assertTrue(bool(recovered["DAT_RECOVERY_USED"]))
        self.assertEqual(recovered["DAT_RECOVERY_SOURCE_DATE"], "2026-06-24")
        self.assertEqual(int(recovered["DAT_RECOVERY_SOURCE_ROW"]), 3)
        self.assertEqual(
            recovered["MATCH_REASON"],
            "STREAM NOT FOUND AFTER ACTUAL MOVEMENT, ORIGINAL, AND RECOVERED DATE SEARCH",
        )
        self.assertEqual(recovered["VALIDASI"], VALIDATION_NOT_FOUND)
        self.assertEqual(recovered["STREAM MATCH MODE"], "NO STREAM MATCH")
        self.assertEqual(len(result["duplicates"]), 1)
        self.assertTrue(bool(result["duplicates"].iloc[0]["USED_FOR_RECOVERY"]))

    def test_ctv010_recovered_date_stream_candidate_is_matched(self):
        dep = pd.DataFrame(columns=["callsign", "adep", "ades", "eobd"])
        arr = pd.DataFrame(
            {
                "callsign": ["CTV010", "CTV010"],
                "adep": ["WIHH", "WIHH"],
                "ades": ["WIMM", "WIMM"],
                "eobd": ["260623", "260624"],
                "eobt": ["2340", "0045"],
                "atd": ["", "2026-06-24 00:54"],
                "ata": ["", "2026-06-24 02:52"],
                "register": ["PKGLM", "PKGLM"],
                "arrivalGate": ["", "27"],
                "arrivalRunway": ["23", "23"],
                "timeStamp": ["2026-06-24 01:49", "2026-06-24 04:52"],
                "messageNum": ["52538", "52582"],
            }
        )
        stream = pd.DataFrame(
            {
                "DATE OF FLIGHT": ["2026-06-23", "2026-06-24"],
                "FLIGHT NUMBER": ["CTV010", "CTV010"],
                "AERODROME": ["WIHH", "WIHH"],
                "TO FROM": ["WIMM", "WIMM"],
                "AC REGISTER": ["PKGLM", "PKGLM"],
                "D/A/L/O": ["A", "A"],
                "ATA": ["2026-06-23 02:52", "2026-06-24 02:52"],
                "STATUS FLIGHT": ["REGULER", "REGULER"],
            }
        )

        result = reconcile_dat_vs_stream(dep, arr, stream)
        self.assertTrue(result["missing"].empty)
        self.assertEqual(result["summary"]["matched"], 1)
        self.assertEqual(result["summary"]["extra_in_stream"], 0)
        matched = result["matched"].iloc[0]
        self.assertEqual(matched["DATE OF FLIGHT"], "2026-06-23")
        self.assertEqual(matched["ACTUAL MOVEMENT DATE"], "2026-06-24")
        self.assertEqual(
            matched["MATCH_REASON"],
            "VALID STREAM MATCH BY ACTUAL MOVEMENT DATE",
        )
        self.assertEqual(
            matched["STREAM VALIDATION RESULT"],
            "VALID STREAM CANDIDATE BY ACTUAL MOVEMENT DATE",
        )
        self.assertEqual(
            matched["STREAM MATCH DATE USED"], "ACTUAL MOVEMENT DATE"
        )
        self.assertEqual(
            matched["STREAM MATCH MODE"], "ACTUAL MOVEMENT DATE MATCH"
        )
        self.assertEqual(matched["VALIDASI"], VALIDATION_FOUND)
        self.assertEqual(matched["STREAM MOVEMENT DATETIME"], "2026-06-24 02:52")
        self.assertEqual(float(matched["TIME DIFFERENCE MINUTES"]), 0.0)
        self.assertTrue(bool(matched["DAT_RECOVERY_USED"]))

    def test_recovered_movement_date_is_valid_on_original_stream_key(self):
        dep = pd.DataFrame(columns=["callsign", "adep", "ades", "eobd"])
        arr = pd.DataFrame(
            {
                "callsign": ["CTV012", "CTV012"],
                "adep": ["WIHH", "WIHH"],
                "ades": ["WIMM", "WIMM"],
                "eobd": ["260623", "260624"],
                "eobt": ["2340", "0045"],
                "atd": ["", "2026-06-24 00:54"],
                "ata": ["", "2026-06-24 02:52"],
                "register": ["PKGLM", "PKGLM"],
            }
        )
        stream = pd.DataFrame(
            {
                "DATE OF FLIGHT": ["2026-06-23"],
                "FLIGHT NUMBER": ["CTV012"],
                "AERODROME": ["WIHH"],
                "TO FROM": ["WIMM"],
                "AC REGISTER": ["PKGLM"],
                "D/A/L/O": ["A"],
                "ATA": ["2026-06-24 02:52"],
                "STATUS FLIGHT": ["REGULER"],
            }
        )
        result = reconcile_dat_vs_stream(dep, arr, stream)
        matched = result["matched"].iloc[0]
        self.assertEqual(
            matched["MATCH_REASON"],
            "VALID STREAM MATCH BY ACTUAL MOVEMENT DATE",
        )
        self.assertEqual(
            matched["STREAM MATCH MODE"], "ACTUAL MOVEMENT DATE MATCH"
        )
        self.assertEqual(float(matched["TIME DIFFERENCE MINUTES"]), 0.0)

    def test_recovered_time_match_with_register_mismatch_needs_review(self):
        dep = pd.DataFrame(columns=["callsign", "adep", "ades", "eobd"])
        arr = pd.DataFrame(
            {
                "callsign": ["CTV013", "CTV013"],
                "adep": ["WIHH", "WIHH"],
                "ades": ["WIMM", "WIMM"],
                "eobd": ["260623", "260624"],
                "eobt": ["2340", "0045"],
                "atd": ["", "2026-06-24 00:54"],
                "ata": ["", "2026-06-24 02:52"],
                "register": ["PKGLM", "PKGLM"],
            }
        )
        stream = pd.DataFrame(
            {
                "DATE OF FLIGHT": ["2026-06-24"],
                "FLIGHT NUMBER": ["CTV013"],
                "AERODROME": ["WIHH"],
                "TO FROM": ["WIMM"],
                "AC REGISTER": ["PKXXX"],
                "D/A/L/O": ["A"],
                "ATA": ["2026-06-24 02:52"],
                "STATUS FLIGHT": ["REGULER"],
            }
        )
        result = reconcile_dat_vs_stream(dep, arr, stream)
        self.assertTrue(result["missing"].empty)
        self.assertTrue(result["matched"].empty)
        reviewed = result["need_review"].iloc[0]
        self.assertEqual(reviewed["MATCH_REASON"], "STREAM AC REGISTER MISMATCH")
        self.assertEqual(
            reviewed["STREAM MATCH MODE"], "ACTUAL MOVEMENT DATE MATCH"
        )
        self.assertEqual(reviewed["VALIDASI"], VALIDATION_REVIEW)
        self.assertEqual(float(reviewed["TIME DIFFERENCE MINUTES"]), 0.0)

    def test_exact_register_within_tolerance_beats_closer_mismatch(self):
        dep = pd.DataFrame(
            {
                "callsign": ["REG001"],
                "adep": ["WIMM"],
                "ades": ["WIII"],
                "eobd": ["260623"],
                "atd": ["10:00"],
                "register": ["PKAAA"],
            }
        )
        arr = pd.DataFrame(columns=["callsign", "adep", "ades", "eobd"])
        stream = pd.DataFrame(
            {
                "DATE OF FLIGHT": ["2026-06-23", "2026-06-23"],
                "FLIGHT NUMBER": ["REG001", "REG001"],
                "AERODROME": ["WIMM", "WIMM"],
                "TO FROM": ["WIII", "WIII"],
                "AC REGISTER": ["PKXXX", "PKAAA"],
                "D/A/L/O": ["D", "D"],
                "ATD": ["2026-06-23 10:00", "2026-06-23 10:01"],
                "STATUS FLIGHT": ["REGULER", "REGULER"],
            }
        )
        result = reconcile_dat_vs_stream(dep, arr, stream)
        matched = result["matched"].iloc[0]
        self.assertEqual(
            matched["MATCH_REASON"],
            "VALID STREAM MATCH BY ACTUAL MOVEMENT DATE",
        )
        self.assertEqual(float(matched["TIME DIFFERENCE MINUTES"]), 1.0)
        self.assertEqual(int(matched["STREAM SOURCE ROW"]), 3)

    def test_actual_date_creates_distinct_flight_instances(self):
        dep = pd.DataFrame(columns=["callsign", "adep", "ades", "eobd"])
        arr = pd.DataFrame(
            {
                "callsign": ["CTV010", "CTV010"],
                "adep": ["WIHH", "WIHH"],
                "ades": ["WIMM", "WIMM"],
                "eobd": ["260624", "260624"],
                "atd": ["2026-06-24 00:54", "2026-06-25 00:09"],
                "ata": ["2026-06-24 02:52", "2026-06-25 02:09"],
                "register": ["PKGLM", "PKGLM"],
                "arrivalGate": ["27", "28"],
                "arrivalRunway": ["23", "23"],
            }
        )
        stream = pd.DataFrame(
            columns=["DATE OF FLIGHT", "FLIGHT NUMBER", "AERODROME", "TO FROM", "D/A/L/O"]
        )
        result = reconcile_dat_vs_stream(dep, arr, stream)
        self.assertEqual(len(result["dat_unique"]), 2)
        self.assertEqual(len(result["duplicates"]), 0)
        self.assertEqual(
            set(result["missing"]["ACTUAL MOVEMENT DATE"]),
            {"2026-06-24", "2026-06-25"},
        )

    def test_overnight_arrival_keeps_eobd_and_actual_movement_date(self):
        dep = pd.DataFrame(columns=["callsign", "adep", "ades", "eobd"])
        arr = pd.DataFrame(
            {
                "callsign": ["CTV010"],
                "adep": ["WIHH"],
                "ades": ["WIMM"],
                "eobd": ["260622"],
                "atd": ["2026-06-22 23:46:00"],
                "ata": ["2026-06-23 01:42:00"],
                "register": ["PKGLM"],
                "arrivalRunway": ["23"],
                "arrivalGate": ["27"],
            }
        )
        stream = pd.DataFrame(
            {
                "DATE OF FLIGHT": ["2026-06-22"],
                "FLIGHT NUMBER": ["CTV010"],
                "AERODROME": ["WIHH"],
                "TO FROM": ["WIMM"],
                "D/A/L/O": ["A"],
                "ATA": ["2026-06-23 01:42:00"],
                "STATUS FLIGHT": ["REGULER"],
            }
        )
        result = reconcile_dat_vs_stream(dep, arr, stream)
        matched = result["matched"].iloc[0]
        self.assertEqual(matched["DATE OF FLIGHT"], "2026-06-22")
        self.assertEqual(matched["ACTUAL MOVEMENT DATE"], "2026-06-23")
        self.assertEqual(matched["ATA"], "01:42")
        self.assertEqual(
            matched["MATCH_REASON"],
            "VALID STREAM MATCH BY ACTUAL MOVEMENT DATE",
        )
        self.assertEqual(
            matched["STREAM MATCH DATE USED"], "ACTUAL MOVEMENT DATE"
        )
        self.assertEqual(
            matched["STREAM MATCH MODE"], "ACTUAL MOVEMENT DATE MATCH"
        )

    def test_overnight_arrival_can_match_by_actual_movement_date(self):
        dep = pd.DataFrame(columns=["callsign", "adep", "ades", "eobd"])
        arr = pd.DataFrame(
            {
                "callsign": ["CTV011"],
                "adep": ["WIHH"],
                "ades": ["WIMM"],
                "eobd": ["260622"],
                "atd": ["2026-06-22 23:46:00"],
                "ata": ["2026-06-23 01:42:00"],
                "register": ["PKGLN"],
                "arrivalRunway": ["23"],
                "arrivalGate": ["27"],
            }
        )
        stream = pd.DataFrame(
            {
                "DATE OF FLIGHT": ["2026-06-23"],
                "FLIGHT NUMBER": ["CTV011"],
                "AERODROME": ["WIHH"],
                "TO FROM": ["WIMM"],
                "AC REGISTER": ["PKGLN"],
                "D/A/L/O": ["A"],
                "ATA": ["2026-06-23 01:42:00"],
                "STATUS FLIGHT": ["REGULER"],
            }
        )
        result = reconcile_dat_vs_stream(dep, arr, stream)
        matched = result["matched"].iloc[0]
        self.assertEqual(
            matched["MATCH_REASON"],
            "VALID STREAM MATCH BY ACTUAL MOVEMENT DATE",
        )
        self.assertEqual(
            matched["STREAM MATCH DATE USED"], "ACTUAL MOVEMENT DATE"
        )
        self.assertEqual(
            matched["STREAM MATCH MODE"], "ACTUAL MOVEMENT DATE MATCH"
        )
        self.assertEqual(result["summary"]["extra_in_stream"], 0)


if __name__ == "__main__":
    unittest.main()

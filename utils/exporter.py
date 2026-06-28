from __future__ import annotations

import io
from datetime import datetime
from typing import Mapping

import pandas as pd

from utils.reconciliation import RECOVERY_AUDIT_COLUMNS, RESULT_COLUMNS


MISSING_EXPORT_COLUMNS = RESULT_COLUMNS + [
    "MATCH_REASON",
    "STREAM STATUS FLIGHT",
    "DAT MOVEMENT DATETIME",
    "STREAM MOVEMENT DATETIME",
    "TIME DIFFERENCE MINUTES",
] + RECOVERY_AUDIT_COLUMNS


def _select_columns(frame: pd.DataFrame, requested: list[str] | None) -> pd.DataFrame:
    if requested is None:
        return frame.copy()
    result = frame.copy()
    for column in requested:
        if column not in result.columns:
            result[column] = ""
    return result[requested]


def build_excel_report(results: Mapping[str, object]) -> bytes:
    """Create the eight-sheet FINDER audit workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    navy = "0B1930"
    blue = "1769E0"
    pale_blue = "EAF2FF"
    white = "FFFFFF"
    border_color = "D7DFEC"
    sheet_colors = {
        "Missing in Stream": "E84C3D",
        "Matched": "1F9D68",
        "Need Review": "F28B30",
        "Extra in Stream": "7567C7",
        "Duplicate DAT": "D99A1B",
        "Audit Detail": "53657D",
        "Excluded Non-Billable": "64748B",
    }

    workbook = Workbook()
    workbook.remove(workbook.active)
    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.sheet_view.showGridLines = False
    summary_sheet.merge_cells("A1:D1")
    summary_sheet["A1"] = "FINDER — DAT vs STREAM Reconciliation Report"
    summary_sheet["A1"].fill = PatternFill("solid", fgColor=navy)
    summary_sheet["A1"].font = Font(color=white, bold=True, size=16)
    summary_sheet["A1"].alignment = Alignment(vertical="center")
    summary_sheet.row_dimensions[1].height = 34
    summary_sheet["A2"] = "Generated"
    summary_sheet["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    settings = results.get("settings", {})
    if isinstance(settings, Mapping):
        summary_sheet["A3"] = "Time tolerance"
        summary_sheet["B3"] = f"{settings.get('time_tolerance_minutes', 30)} minutes"
        summary_sheet["C3"] = "Invalid STREAM status"
        summary_sheet["D3"] = ", ".join(settings.get("invalid_stream_statuses", ["OTHER"]))
    summary_sheet["A4"] = "METRIC"
    summary_sheet["B4"] = "VALUE"
    for cell in summary_sheet[4][:2]:
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(color=white, bold=True)

    summary_table = results["summary_table"]
    assert isinstance(summary_table, pd.DataFrame)
    for row_index, row in enumerate(summary_table.itertuples(index=False), start=5):
        summary_sheet.cell(row_index, 1, row.METRIC)
        value_cell = summary_sheet.cell(row_index, 2, row.VALUE)
        if row.METRIC == "Accuracy Percentage":
            value_cell.number_format = "0.0%"
        if row.METRIC == "Missing in Stream":
            for column in (1, 2):
                summary_sheet.cell(row_index, column).fill = PatternFill(
                    "solid", fgColor="FDE9E7"
                )
                summary_sheet.cell(row_index, column).font = Font(
                    color="B42318", bold=True
                )
        elif row.METRIC == "Need Review":
            for column in (1, 2):
                summary_sheet.cell(row_index, column).fill = PatternFill(
                    "solid", fgColor="FFF1E5"
                )
        elif row_index % 2:
            for column in (1, 2):
                summary_sheet.cell(row_index, column).fill = PatternFill(
                    "solid", fgColor=pale_blue
                )
    summary_sheet.column_dimensions["A"].width = 38
    summary_sheet.column_dimensions["B"].width = 22
    summary_sheet.column_dimensions["C"].width = 24
    summary_sheet.column_dimensions["D"].width = 28
    summary_sheet.freeze_panes = "A5"

    sheet_specs = [
        ("Missing in Stream", "missing", MISSING_EXPORT_COLUMNS),
        ("Matched", "matched", None),
        ("Need Review", "need_review", None),
        ("Extra in Stream", "extra", None),
        ("Duplicate DAT", "duplicates", None),
        ("Audit Detail", "audit_detail", None),
        ("Excluded Non-Billable", "excluded_non_billable", None),
    ]
    thin = Side(style="thin", color=border_color)
    text_columns = {
        "FLIGHT NUMBER",
        "AERODROME",
        "TO FROM",
        "AC REGISTER",
        "D/A/L/O",
        "ARRIVAL GATE",
        "DEPARTURE GATE",
        "DEPARTURE RUNWAY",
        "ARRIVAL RUNWAY",
        "SOURCE ROW",
    }
    wrapped_columns = {
        "MATCH_REASON",
        "DUPLICATE_REASON",
        "DAT_RECOVERY_REASON",
        "STREAM VALIDATION RESULT",
        "FLIGHT_INSTANCE_KEY",
        "DUPLICATE_GROUP_KEY",
    }
    for sheet_name, result_key, requested_columns in sheet_specs:
        raw_frame = results[result_key]
        assert isinstance(raw_frame, pd.DataFrame)
        frame = _select_columns(raw_frame, requested_columns)
        sheet = workbook.create_sheet(sheet_name)
        sheet.sheet_view.showGridLines = False
        last_column = get_column_letter(max(1, len(frame.columns)))
        sheet.merge_cells(f"A1:{last_column}1")
        sheet["A1"] = sheet_name.upper()
        sheet["A1"].fill = PatternFill("solid", fgColor=sheet_colors[sheet_name])
        sheet["A1"].font = Font(color=white, bold=True, size=14)
        sheet["A1"].alignment = Alignment(vertical="center")
        sheet.row_dimensions[1].height = 30
        sheet["A2"] = f"Rows: {len(frame):,}"

        header_row = 4
        for column_index, column in enumerate(frame.columns, start=1):
            cell = sheet.cell(header_row, column_index, str(column))
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(color=white, bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)

        for row_index, values in enumerate(
            frame.itertuples(index=False, name=None), start=5
        ):
            longest_wrapped_value = 0
            for column_index, value in enumerate(values, start=1):
                column_name = frame.columns[column_index - 1]
                clean_value = "" if pd.isna(value) else value
                if column_name in {
                    "DATE OF FLIGHT",
                    "ACTUAL MOVEMENT DATE",
                    "DAT_RECOVERY_SOURCE_DATE",
                    "ORIGINAL_DAT_DATE",
                    "RECOVERED_DAT_DATE",
                } and clean_value:
                    try:
                        clean_value = datetime.strptime(
                            str(clean_value), "%Y-%m-%d"
                        ).date()
                    except ValueError:
                        pass
                cell = sheet.cell(row_index, column_index, clean_value)
                cell.alignment = Alignment(
                    vertical="top", wrap_text=column_name in wrapped_columns
                )
                if column_name in wrapped_columns:
                    longest_wrapped_value = max(
                        longest_wrapped_value, len(str(clean_value))
                    )
                if column_name in {
                    "DATE OF FLIGHT",
                    "ACTUAL MOVEMENT DATE",
                    "DAT_RECOVERY_SOURCE_DATE",
                    "ORIGINAL_DAT_DATE",
                    "RECOVERED_DAT_DATE",
                }:
                    cell.number_format = "yyyy-mm-dd"
                elif column_name in text_columns:
                    cell.number_format = "@"
                elif column_name == "TIME DIFFERENCE MINUTES" and clean_value != "":
                    cell.number_format = "0.0"
                if row_index % 2:
                    cell.fill = PatternFill("solid", fgColor="F7F9FC")
            if longest_wrapped_value > 38:
                estimated_lines = (longest_wrapped_value + 37) // 38
                sheet.row_dimensions[row_index].height = min(60, 15 * estimated_lines)

        if len(frame):
            sheet.auto_filter.ref = f"A4:{last_column}{len(frame) + 4}"
        sheet.freeze_panes = "A5"
        sheet.row_dimensions[4].height = 36
        for column_index, column in enumerate(frame.columns, start=1):
            sample = [str(column)] + [
                str(value)
                for value in frame[column].head(200).tolist()
                if not pd.isna(value)
            ]
            width = min(max(len(value) for value in sample) + 2, 34)
            sheet.column_dimensions[get_column_letter(column_index)].width = max(
                width, 11
            )
            if column in wrapped_columns:
                sheet.column_dimensions[get_column_letter(column_index)].width = 38

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()

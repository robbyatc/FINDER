from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path
from typing import Mapping

import pandas as pd

from comparison import (
    actual_file_format,
    actual_file_format_label,
    canonicalize,
    detect_mapping,
    excel_sheet_names,
    normalize_value,
    read_actual_excel,
    read_source_csv,
)


RESULT_COLUMNS = [
    "DATE OF FLIGHT",
    "FLIGHT NUMBER",
    "AERODROME",
    "TO FROM",
    "AC REGISTER",
    "ATD",
    "ATA",
    "D/A/L/O",
    "ARRIVAL GATE",
    "DEPARTURE GATE",
    "DEPARTURE RUNWAY",
    "ARRIVAL RUNWAY",
]

DETAIL_COLUMNS = RESULT_COLUMNS + ["SOURCE DATA", "SOURCE ROW", "STATUS"]

MATCH_KEY_SPECS = [
    ("DATE OF FLIGHT", "eobd"),
    ("FLIGHT NUMBER", "flight"),
    ("AERODROME", "adep"),
    ("TO FROM", "ades"),
    ("D/A/L/O", "movement"),
]

REQUIRED_DAT_FIELDS = ["flight", "eobd", "adep", "ades"]
REQUIRED_STREAM_FIELDS = ["flight", "eobd", "adep", "ades", "movement"]


def read_uploaded_table(data: bytes, filename: str) -> tuple[pd.DataFrame, str]:
    """Read CSV, XLS, XLSX, or an HTML report carrying an .xls extension."""
    suffix = Path(filename).suffix.lower()
    physical_format = actual_file_format(data)

    if suffix in {".csv", ".tsv", ".txt"} or physical_format == "unknown":
        try:
            frame, encoding = read_source_csv(data)
            return frame, f"Delimited text · {encoding}"
        except Exception:
            if physical_format == "unknown":
                raise

    sheets = excel_sheet_names(data)
    frame = read_actual_excel(data, sheets[0])
    return frame, f"{actual_file_format_label(data)} · {sheets[0]}"


def detected_mapping(frame: pd.DataFrame) -> dict[str, str | None]:
    return detect_mapping(frame.columns)


def validate_required_columns(
    dep_mapping: Mapping[str, str | None],
    arr_mapping: Mapping[str, str | None],
    stream_mapping: Mapping[str, str | None],
) -> list[str]:
    issues: list[str] = []
    labels = {
        "flight": "Flight Number/Callsign",
        "eobd": "Date of Flight/EOBD",
        "adep": "ADEP/Aerodrome",
        "ades": "ADES/TO FROM",
        "movement": "D/A/L/O",
    }
    for dataset, mapping, required in (
        ("DAT DEP", dep_mapping, REQUIRED_DAT_FIELDS),
        ("DAT ARR", arr_mapping, REQUIRED_DAT_FIELDS),
        ("STREAM", stream_mapping, REQUIRED_STREAM_FIELDS),
    ):
        missing = [labels[field] for field in required if not mapping.get(field)]
        if missing:
            issues.append(f"{dataset}: {', '.join(missing)}")
    return issues


def _present(value: object) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    return text not in {"", "-", "—", "N/A", "NA", "NULL", "NONE"}


def _prefer(primary: object, fallback: object) -> str:
    return str(primary) if _present(primary) else (str(fallback) if _present(fallback) else "")


def standardize_dataset(
    frame: pd.DataFrame,
    mapping: Mapping[str, str | None],
    source_name: str,
    default_movement: str | None = None,
) -> pd.DataFrame:
    canonical = canonicalize(frame, mapping)
    if default_movement:
        movement = pd.Series(default_movement, index=canonical.index, dtype=object)
    else:
        movement = canonical["movement"].map(lambda value: normalize_value(value, "movement"))

    records: list[dict[str, object]] = []
    for index, row in canonical.iterrows():
        move = str(movement.loc[index]).upper().strip()
        generic_gate = row["parking"]
        generic_runway = row["runway"]
        arrival_gate = row["arrival_gate"]
        departure_gate = row["departure_gate"]
        arrival_runway = row["arrival_runway"]
        departure_runway = row["departure_runway"]

        if move == "A":
            arrival_gate = _prefer(arrival_gate, generic_gate)
            arrival_runway = _prefer(arrival_runway, generic_runway)
        elif move == "D":
            departure_gate = _prefer(departure_gate, generic_gate)
            departure_runway = _prefer(departure_runway, generic_runway)

        records.append(
            {
                "DATE OF FLIGHT": row["eobd"],
                "FLIGHT NUMBER": row["flight"],
                "AERODROME": row["adep"],
                "TO FROM": row["ades"],
                "AC REGISTER": row["register"],
                "ATD": row["atd"],
                "ATA": row["ata"],
                "D/A/L/O": move,
                "ARRIVAL GATE": arrival_gate,
                "DEPARTURE GATE": departure_gate,
                "DEPARTURE RUNWAY": departure_runway,
                "ARRIVAL RUNWAY": arrival_runway,
                "SOURCE DATA": source_name,
                "SOURCE ROW": int(row["__row"]),
            }
        )
    return pd.DataFrame(records, columns=DETAIL_COLUMNS[:-1])


def _prepare_keys(frame: pd.DataFrame, side: str) -> pd.DataFrame:
    prepared = frame.copy().reset_index(drop=True)
    key_columns: list[str] = []
    for column, field in MATCH_KEY_SPECS:
        key_column = f"__key_{field}"
        prepared[key_column] = prepared[column].map(
            lambda value, canonical_field=field: normalize_value(value, canonical_field)
        )
        key_columns.append(key_column)

    prepared["__complete_key"] = prepared[key_columns].ne("").all(axis=1)
    prepared["__occurrence"] = 0
    complete = prepared["__complete_key"]
    prepared.loc[complete, "__occurrence"] = prepared.loc[complete].groupby(
        key_columns, dropna=False, sort=False
    ).cumcount()
    prepared.loc[complete, "__match_id"] = (
        prepared.loc[complete, key_columns].astype(str).agg("␟".join, axis=1)
        + "␟"
        + prepared.loc[complete, "__occurrence"].astype(str)
    )
    prepared.loc[~complete, "__match_id"] = prepared.loc[~complete].apply(
        lambda row: f"__INCOMPLETE__{side}__{row['SOURCE DATA']}__{row['SOURCE ROW']}",
        axis=1,
    )
    return prepared


def _clean_result(frame: pd.DataFrame, status: str) -> pd.DataFrame:
    result = frame[DETAIL_COLUMNS[:-1]].copy()
    result["STATUS"] = status
    if not result.empty:
        result = result.sort_values(
            ["DATE OF FLIGHT", "FLIGHT NUMBER", "D/A/L/O"],
            kind="stable",
        ).reset_index(drop=True)
    return result[DETAIL_COLUMNS]


def _merge_matched(matched: pd.DataFrame) -> pd.DataFrame:
    records: list[dict[str, object]] = []
    for _, row in matched.iterrows():
        record: dict[str, object] = {}
        for column in RESULT_COLUMNS:
            record[column] = _prefer(row[f"dat_{column}"], row[f"stream_{column}"])
        record["SOURCE DATA"] = row["dat_SOURCE DATA"]
        record["SOURCE ROW"] = int(row["dat_SOURCE ROW"])
        record["STATUS"] = "MATCHED"
        records.append(record)
    return pd.DataFrame(records, columns=DETAIL_COLUMNS)


def reconcile_dat_vs_stream(
    dat_dep: pd.DataFrame,
    dat_arr: pd.DataFrame,
    stream: pd.DataFrame,
    dep_mapping: Mapping[str, str | None] | None = None,
    arr_mapping: Mapping[str, str | None] | None = None,
    stream_mapping: Mapping[str, str | None] | None = None,
) -> dict[str, object]:
    dep_mapping = dict(dep_mapping or detected_mapping(dat_dep))
    arr_mapping = dict(arr_mapping or detected_mapping(dat_arr))
    stream_mapping = dict(stream_mapping or detected_mapping(stream))
    issues = validate_required_columns(dep_mapping, arr_mapping, stream_mapping)
    if issues:
        raise ValueError("Kolom wajib belum lengkap: " + " | ".join(issues))

    dep = standardize_dataset(dat_dep, dep_mapping, "DAT DEP", "D")
    arr = standardize_dataset(dat_arr, arr_mapping, "DAT ARR", "A")
    stream_standard = standardize_dataset(stream, stream_mapping, "STREAM")
    dat_combined = pd.concat([dep, arr], ignore_index=True)

    dat_prepared = _prepare_keys(dat_combined, "DAT")
    stream_prepared = _prepare_keys(stream_standard, "STREAM")
    key_columns = [f"__key_{field}" for _, field in MATCH_KEY_SPECS]

    duplicate_mask = dat_prepared["__complete_key"] & dat_prepared.duplicated(
        key_columns, keep="first"
    )
    duplicate_dat = _clean_result(
        dat_prepared.loc[duplicate_mask], "DUPLICATE DAT"
    )
    dat_unique = dat_prepared.loc[~duplicate_mask].copy()
    # Rebuild occurrence after DAT duplicates have been removed.
    dat_unique = _prepare_keys(dat_unique[DETAIL_COLUMNS[:-1]], "DAT")

    dat_ids = set(dat_unique["__match_id"])
    stream_ids = set(stream_prepared["__match_id"])
    shared_ids = dat_ids & stream_ids

    missing = _clean_result(
        dat_unique.loc[~dat_unique["__match_id"].isin(stream_ids)],
        "MISSING IN STREAM",
    )
    extra = _clean_result(
        stream_prepared.loc[~stream_prepared["__match_id"].isin(dat_ids)],
        "EXTRA IN STREAM",
    )

    dat_match = dat_unique.loc[dat_unique["__match_id"].isin(shared_ids)]
    stream_match = stream_prepared.loc[stream_prepared["__match_id"].isin(shared_ids)]
    dat_columns = DETAIL_COLUMNS[:-1] + ["__match_id"]
    stream_columns = DETAIL_COLUMNS[:-1] + ["__match_id"]
    matched_raw = dat_match[dat_columns].rename(
        columns={column: f"dat_{column}" for column in DETAIL_COLUMNS[:-1]}
    ).merge(
        stream_match[stream_columns].rename(
            columns={column: f"stream_{column}" for column in DETAIL_COLUMNS[:-1]}
        ),
        on="__match_id",
        how="inner",
        validate="one_to_one",
    )
    matched = _merge_matched(matched_raw)
    if not matched.empty:
        matched = matched.sort_values(
            ["DATE OF FLIGHT", "FLIGHT NUMBER", "D/A/L/O"], kind="stable"
        ).reset_index(drop=True)

    unique_dat_total = len(dat_unique)
    matched_total = len(matched)
    accuracy = (matched_total / unique_dat_total * 100) if unique_dat_total else 0.0
    summary = {
        "total_dat_dep": len(dep),
        "total_dat_arr": len(arr),
        "total_dat_combined": len(dat_combined),
        "total_dat_unique": unique_dat_total,
        "total_stream": len(stream_standard),
        "matched": matched_total,
        "missing_in_stream": len(missing),
        "extra_in_stream": len(extra),
        "duplicate_dat": len(duplicate_dat),
        "accuracy_percentage": accuracy,
        "incomplete_dat_keys": int((~dat_unique["__complete_key"]).sum()),
        "incomplete_stream_keys": int((~stream_prepared["__complete_key"]).sum()),
    }
    summary_table = pd.DataFrame(
        [
            ("Total DAT DEP", summary["total_dat_dep"]),
            ("Total DAT ARR", summary["total_dat_arr"]),
            ("Total DAT Combined", summary["total_dat_combined"]),
            ("Total DAT Unique", summary["total_dat_unique"]),
            ("Total STREAM", summary["total_stream"]),
            ("Matched", summary["matched"]),
            ("Missing in Stream", summary["missing_in_stream"]),
            ("Extra in Stream", summary["extra_in_stream"]),
            ("Duplicate DAT", summary["duplicate_dat"]),
            ("Accuracy Percentage", summary["accuracy_percentage"] / 100),
        ],
        columns=["METRIC", "VALUE"],
    )
    return {
        "summary": summary,
        "summary_table": summary_table,
        "missing": missing,
        "matched": matched,
        "extra": extra,
        "duplicates": duplicate_dat,
        "dat_dep": dep,
        "dat_arr": arr,
        "stream": stream_standard,
    }


def build_excel_report(results: Mapping[str, object]) -> bytes:
    """Create a styled five-sheet Excel report for the Streamlit download button."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    navy = "0B1930"
    blue = "1769E0"
    pale_blue = "EAF2FF"
    white = "FFFFFF"
    border_color = "D7DFEC"
    sheet_colors = {
        "Missing in Stream": "E84C3D",
        "Matched": "1F9D68",
        "Extra in Stream": "7567C7",
        "Duplicate DAT": "D99A1B",
    }

    workbook = Workbook()
    workbook.remove(workbook.active)

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.sheet_view.showGridLines = False
    summary_sheet.merge_cells("A1:D1")
    summary_sheet["A1"] = "FINDER — DAT vs STREAM Reconciliation Report"
    summary_sheet["A1"].fill = PatternFill("solid", fgColor=navy)
    summary_sheet["A1"].font = Font(color=white, bold=True, size=16)
    summary_sheet["A1"].alignment = Alignment(vertical="center")
    summary_sheet.row_dimensions[1].height = 34
    summary_sheet["A2"] = "Generated"
    summary_sheet["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary_sheet["A4"] = "METRIC"
    summary_sheet["B4"] = "VALUE"
    for cell in summary_sheet[4][:2]:
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(color=white, bold=True)

    summary_table = results["summary_table"]
    assert isinstance(summary_table, pd.DataFrame)
    for row_index, row in enumerate(summary_table.itertuples(index=False), start=5):
        summary_sheet.cell(row_index, 1, row.METRIC)
        value_cell = summary_sheet.cell(row_index, 2, row.VALUE)
        if row.METRIC == "Accuracy Percentage":
            value_cell.number_format = "0.0%"
        if row.METRIC == "Missing in Stream":
            for column in (1, 2):
                summary_sheet.cell(row_index, column).fill = PatternFill(
                    "solid", fgColor="FDE9E7"
                )
                summary_sheet.cell(row_index, column).font = Font(
                    color="B42318", bold=True
                )
        elif row_index % 2:
            for column in (1, 2):
                summary_sheet.cell(row_index, column).fill = PatternFill(
                    "solid", fgColor=pale_blue
                )
    summary_sheet.column_dimensions["A"].width = 28
    summary_sheet.column_dimensions["B"].width = 20
    summary_sheet.freeze_panes = "A5"

    thin = Side(style="thin", color=border_color)
    for sheet_name, result_key in (
        ("Missing in Stream", "missing"),
        ("Matched", "matched"),
        ("Extra in Stream", "extra"),
        ("Duplicate DAT", "duplicates"),
    ):
        frame = results[result_key]
        assert isinstance(frame, pd.DataFrame)
        sheet = workbook.create_sheet(sheet_name)
        sheet.sheet_view.showGridLines = False
        last_column = get_column_letter(max(1, len(frame.columns)))
        sheet.merge_cells(f"A1:{last_column}1")
        sheet["A1"] = sheet_name.upper()
        sheet["A1"].fill = PatternFill("solid", fgColor=sheet_colors[sheet_name])
        sheet["A1"].font = Font(color=white, bold=True, size=14)
        sheet["A1"].alignment = Alignment(vertical="center")
        sheet.row_dimensions[1].height = 30
        sheet["A2"] = f"Rows: {len(frame):,}"

        header_row = 4
        for column_index, column in enumerate(frame.columns, start=1):
            cell = sheet.cell(header_row, column_index, str(column))
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(color=white, bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)

        for row_index, values in enumerate(frame.itertuples(index=False, name=None), start=5):
            for column_index, value in enumerate(values, start=1):
                clean_value = "" if pd.isna(value) else value
                cell = sheet.cell(row_index, column_index, clean_value)
                cell.alignment = Alignment(vertical="top")
                if frame.columns[column_index - 1] in {
                    "FLIGHT NUMBER",
                    "AERODROME",
                    "TO FROM",
                    "AC REGISTER",
                    "D/A/L/O",
                    "ARRIVAL GATE",
                    "DEPARTURE GATE",
                    "DEPARTURE RUNWAY",
                    "ARRIVAL RUNWAY",
                }:
                    cell.number_format = "@"
                if row_index % 2:
                    cell.fill = PatternFill("solid", fgColor="F7F9FC")

        if len(frame):
            sheet.auto_filter.ref = f"A4:{last_column}{len(frame) + 4}"
        sheet.freeze_panes = "A5"
        sheet.row_dimensions[4].height = 28
        for column_index, column in enumerate(frame.columns, start=1):
            sample = [str(column)] + [
                str(value)
                for value in frame[column].head(200).tolist()
                if not pd.isna(value)
            ]
            width = min(max(len(value) for value in sample) + 2, 28)
            sheet.column_dimensions[get_column_letter(column_index)].width = max(width, 11)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
